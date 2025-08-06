import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import tempfile
import yagmail
from copy import copy

# Function to copy formatting from one row to another
def copy_row_formatting(ws: Worksheet, source_row: int, target_row: int):
    for col in range(1, ws.max_column + 1):
        cell_src = ws.cell(row=source_row, column=col)
        cell_tgt = ws.cell(row=target_row, column=col)

        if cell_src.has_style:
            cell_tgt.font = copy(cell_src.font)
            cell_tgt.border = copy(cell_src.border)
            cell_tgt.fill = copy(cell_src.fill)
            cell_tgt.number_format = copy(cell_src.number_format)
            cell_tgt.alignment = copy(cell_src.alignment)
            cell_tgt.protection = copy(cell_src.protection)

# Streamlit page settings
st.set_page_config(page_title="Project Report Generator", layout="wide")
st.title("📊 Automated Project Report Generator")

# Store PM items in session
if "pm_items" not in st.session_state:
    st.session_state.pm_items = []

# Step 1: Basic Details
st.header("Step 1: Basic Details")
firm_name = st.text_input("Firm Name").upper()
firm_address = st.text_input("Firm Address").upper()
nature_of_business = st.text_input("Nature of Business").upper()
proprietor_name = st.text_input("Proprietor Name").upper()
proprietor_address = st.text_input("Address of Proprietor").upper()
building_status = st.selectbox("Building Rented/Owned", ["RENTED", "OWNED"])
area_sqft = st.text_input("Area in sq. ft").upper()
rent_rs = st.text_input("Rent in Rs.") if building_status == "RENTED" else ""

# Step 2: Financial Parameters
st.header("Step 2: Financial Parameters")
margin_percent = st.number_input("Margin Percentage", min_value=0.0, max_value=100.0)
subsidy_percent = st.number_input("Subsidy Percentage", min_value=0.0, max_value=100.0)
term_loan_interest = st.number_input("Term Loan Interest Rate", min_value=0.0, max_value=100.0)
cc_interest = st.number_input("CC Interest Rate", min_value=0.0, max_value=100.0)

# Decimal conversions
margin_decimal = margin_percent / 100
subsidy_decimal = subsidy_percent / 100
tli_decimal = term_loan_interest / 100
cci_decimal = cc_interest / 100

# Step 3: Add Plant & Machinery Items
st.header("Step 3: Plant & Machinery")
with st.form("pm_form", clear_on_submit=True):
    col1, col2, col3 = st.columns([4, 2, 2])
    with col1:
        pm_name = st.text_input("Item Name")
    with col2:
        pm_qty = st.number_input("Quantity", min_value=0.0)
    with col3:
        pm_rate = st.number_input("Rate", min_value=0.0)
    submitted = st.form_submit_button("Add Item")
    if submitted and pm_name:
        st.session_state.pm_items.append({
            "name": pm_name,
            "qty": pm_qty,
            "rate": pm_rate,
            "total": pm_qty * pm_rate
        })

if st.session_state.pm_items:
    st.subheader("Items Added")
    st.table([
        {
            "S.No": i + 1,
            "Item Name": item["name"],
            "Quantity": item["qty"],
            "Rate": item["rate"],
            "Total": item["total"]
        }
        for i, item in enumerate(st.session_state.pm_items)
    ])

# Step 4: Furniture and Electrical
st.header("Step 4: Other Fixed Assets")
furniture_value = st.number_input("Furniture & Fixture Value", min_value=0.0)
electrical_value = st.number_input("Electrical Equipments & Fittings Value", min_value=0.0)

# Summary
pm_total = sum(item["total"] for item in st.session_state.pm_items)
grand_total = pm_total + furniture_value + electrical_value

st.markdown("### 💰 Summary")
st.write(f"**Total Plant & Machinery:** ₹{pm_total:,.2f}")
st.write(f"**Furniture & Fixtures:** ₹{furniture_value:,.2f}")
st.write(f"**Electrical Equipments:** ₹{electrical_value:,.2f}")
st.success(f"**Grand Total:** ₹{grand_total:,.2f}")

# Generate and Email Report
if st.button("Generate Project Report"):
    template_path = "Project Report Format.xlsx"
    wb = load_workbook(template_path)

    # Basic Details
    if "Basic Details" in wb.sheetnames:
        sheet = wb["Basic Details"]
        sheet["C3"] = firm_name
        sheet["C4"] = firm_address
        sheet["C5"] = nature_of_business
        sheet["C6"] = proprietor_name
        sheet["C7"] = proprietor_address
        sheet["C8"] = building_status
        sheet["C9"] = area_sqft
        sheet["C10"] = rent_rs
        sheet["C11"] = margin_decimal
        sheet["C12"] = subsidy_decimal
        sheet["C13"] = tli_decimal
        sheet["C14"] = cci_decimal

    # PM-MFA Sheet
    if "PM-MFA" in wb.sheetnames:
        sheet = wb["PM-MFA"]
        base_row = 10
        default_rows = 1
        n = len(st.session_state.pm_items)

        # Save and restore merged cell ranges
        merged_ranges = list(sheet.merged_cells.ranges)
        if n > default_rows:
            sheet.insert_rows(base_row + 1, n - default_rows)
        for merge in merged_ranges:
            sheet.merge_cells(str(merge))

        # Write P&M Items
        for i, item in enumerate(st.session_state.pm_items):
            row = base_row + i
            copy_row_formatting(sheet, base_row, row)
            sheet[f"A{row}"] = i + 1
            sheet[f"B{row}"] = item["name"]
            sheet[f"C{row}"] = item["qty"]
            sheet[f"D{row}"] = item["rate"]
            sheet[f"E{row}"] = f"=C{row}*D{row}"

        # Add Total Formula
        total_row = base_row + n + 3
        sheet[f"C{total_row}"] = f"=SUM(C{base_row}:C{base_row + n - 1})"

        # Insert Furniture and Electrical
        sheet[f"E{27 + (n - 1)}"] = furniture_value
        sheet[f"E{30 + (n - 1)}"] = electrical_value

        # Fix downstream reference (example: C34 originally refers to C33)
        cell_c34 = sheet["C34"]
        if isinstance(cell_c34.value, str) and "=C33" in cell_c34.value:
            cell_c34.value = f"=C{total_row}"

    # Save and email
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    try:
        st.info("📧 Sending report to Lavish Gupta...")
        yag = yagmail.SMTP(user="calavishgupta25@gmail.com", password="geli tejz dxiq vtyo")
        yag.send(
            to="calavishgupta25@gmail.com",
            subject="New Project Report Submission",
            contents="Attached is the new project report submitted via Streamlit app.",
            attachments=tmp_path
        )
        st.success("✅ Report generated and emailed successfully.")
    except Exception as e:
        st.error(f"❌ Failed to send email: {e}")
